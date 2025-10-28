import os
import io
import secrets
import pickle
import base64
import pandas as pd
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from functools import wraps

from flask import Flask, render_template, Response, jsonify, request, redirect, url_for, flash, session
import pymysql
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET_KEY")

# Database configuration for Railway
def get_db():
    db_config = {
        "host": os.environ.get("MYSQLHOST"),
        "user": os.environ.get("MYSQLUSER", "root"),
        "password": os.environ.get("MYSQLPASSWORD"),
        "database": os.environ.get("MYSQLDATABASE", "railway"),
        "port": int(os.environ.get("MYSQLPORT", 3306)),
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor,
        "autocommit": False,
        "ssl": {"ssl": {"ca": "/etc/ssl/certs/ca-certificates.crt"}}
    }
    return pymysql.connect(**db_config)

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "admin_id" not in session:
            return redirect(url_for("login", next=request.path))
        session.permanent = True
        return f(*args, **kwargs)
    return wrapper

def query_one(sql, params=None):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            return cur.fetchone()
    finally:
        conn.close()

def query_all(sql, params=None):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            return cur.fetchall()
    finally:
        conn.close()

def execute(sql, params=None, many=False):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            if many:
                cur.executemany(sql, params or [])
            else:
                cur.execute(sql, params or ())
        conn.commit()
    finally:
        conn.close()

# ----- Routes -----
@app.route("/")
def index():
    if "admin_id" in session:
        return redirect(url_for("admin_panel"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        row = query_one(
            "SELECT id, name, username, password_hash FROM admin WHERE username=%s",
            (username,),
        )
        if not row or not check_password_hash(row["password_hash"], password):
            flash("Invalid username/password")
            return render_template("login.html")

        session.permanent = True
        session["admin_id"] = row["id"]
        session["admin_name"] = row["name"]
        return redirect(url_for("admin_panel"))
    return render_template("login.html")

@app.route("/menu")
@login_required
def menu():
    return render_template("menu.html", admin_name=session.get("admin_name", "Admin"))

@app.route("/enroll/<token>", methods=["GET"])
def enroll_get(token):
    row = query_one(
        "SELECT token, created_at_utc FROM pending_enrollments WHERE token=%s", (token,)
    )
    if not row:
        return "Token not found", 404

    created_at = row["created_at_utc"]
    if isinstance(created_at, datetime) and created_at.tzinfo is None:
        created_at = created_at.replace(tzinfo=timezone.utc)

    age = datetime.now(timezone.utc) - created_at
    if age.total_seconds() > 660:  # 10 minutes + grace
        execute("DELETE FROM pending_enrollments WHERE token=%s", (token,))
        return "Token expired", 410

    return render_template("enroll.html", token=token)

@app.route("/enroll/<token>", methods=["POST"])
def enroll_post(token):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT face_encoding, created_at_utc FROM pending_enrollments WHERE token=%s",
                (token,),
            )
            row = cur.fetchone()
            if not row:
                return "Token not found", 404

            created_at = row["created_at_utc"]
            if isinstance(created_at, datetime) and created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)

            age = datetime.now(timezone.utc) - created_at
            if age.total_seconds() > 660:
                cur.execute("DELETE FROM pending_enrollments WHERE token=%s", (token,))
                conn.commit()
                return "Token expired", 410

        name = request.form.get("name", "").strip()
        student_code = request.form.get("student_id", "").strip()
        if not name or not student_code:
            flash("Name and Student ID required")
            return redirect(url_for("enroll_get", token=token))

        with conn.cursor() as cur:
            cur.execute("SELECT id FROM students WHERE student_id=%s", (student_code,))
            existing = cur.fetchone()

            if existing:
                sid = existing["id"]
                cur.execute(
                    "UPDATE students SET name=%s, face_encoding=%s WHERE id=%s",
                    (name, row["face_encoding"], sid),
                )
            else:
                cur.execute(
                    "INSERT INTO students (student_id, name, face_encoding, created_at_utc) VALUES (%s,%s,%s,%s)",
                    (student_code, name, row["face_encoding"], datetime.now(timezone.utc)),
                )
                sid = cur.lastrowid

            cur.execute(
                "INSERT INTO attendance_logs (student_id, type, created_at_utc) VALUES (%s,%s,%s)",
                (sid, "check_in", datetime.now(timezone.utc)),
            )

            cur.execute("DELETE FROM pending_enrollments WHERE token=%s", (token,))

        conn.commit()
        return render_template("success.html", name=name, student_id=student_code)

    finally:
        conn.close()

@app.route("/admin")
@login_required
def admin_panel():
    conn = get_db()
    try:
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        student_filter = request.args.get('student_filter', '')
        type_filter = request.args.get('type_filter', '')
        sort_by = request.args.get('sort_by', 'created_at_utc')
        sort_order = request.args.get('sort_order', 'desc')
        
        attendance_query = """
            SELECT 
                a.id,
                a.type,
                a.created_at_utc,
                s.student_id,
                s.name AS student_name,
                DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) as date_local,
                TIME(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) as time_local
            FROM attendance_logs a
            LEFT JOIN students s ON s.id = a.student_id
            WHERE 1=1
        """
        params = []
        
        if date_from:
            attendance_query += " AND DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) >= %s"
            params.append(date_from)
        if date_to:
            attendance_query += " AND DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) <= %s"
            params.append(date_to)
        if student_filter:
            attendance_query += " AND s.id = %s"
            params.append(student_filter)
        if type_filter:
            attendance_query += " AND a.type = %s"
            params.append(type_filter)
        
        if sort_by == 'student_name':
            sort_column = 's.name'
        elif sort_by == 'student_id':
            sort_column = 's.student_id'
        elif sort_by == 'type':
            sort_column = 'a.type'
        else:
            sort_column = 'a.created_at_utc'
        
        attendance_query += f" ORDER BY {sort_column} {sort_order.upper()} LIMIT 200"
        
        with conn.cursor() as cur:
            cur.execute(attendance_query, params)
            attendance = cur.fetchall()

            cur.execute(
                "SELECT token, face_image, created_at_utc, created_at_display FROM pending_enrollments ORDER BY created_at_utc DESC LIMIT 50"
            )
            pendings = cur.fetchall()

            cur.execute("SELECT id, student_id, name FROM students ORDER BY name ASC")
            students = cur.fetchall()

        return render_template(
            "admin.html",
            admin_name=session.get("admin_name", ""),
            attendance=attendance,
            pendings=pendings,
            students=students,
            current_filters=request.args
        )
    finally:
        conn.close()

@app.route("/admin/face_image/<token>")
@login_required
def get_face_image(token):
    row = query_one("SELECT face_image FROM pending_enrollments WHERE token=%s", (token,))
    if not row or not row["face_image"]:
        return "Image not found", 404
    return Response(row["face_image"], mimetype="image/jpeg")

@app.route("/admin/export_excel")
@login_required
def export_excel():
    try:
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        student_filter = request.args.get('student_filter', '')
        type_filter = request.args.get('type_filter', '')
        sort_by = request.args.get('sort_by', 'created_at_utc')
        sort_order = request.args.get('sort_order', 'desc')
        
        query = """
            SELECT 
                a.id,
                s.student_id,
                s.name AS student_name,
                a.type,
                a.created_at_utc,
                DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) as date_local,
                TIME(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) as time_local
            FROM attendance_logs a
            LEFT JOIN students s ON s.id = a.student_id
            WHERE 1=1
        """
        params = []
        
        if date_from:
            query += " AND DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) >= %s"
            params.append(date_from)
        if date_to:
            query += " AND DATE(CONVERT_TZ(a.created_at_utc, '+00:00', '+08:00')) <= %s"
            params.append(date_to)
        if student_filter:
            query += " AND s.id = %s"
            params.append(student_filter)
        if type_filter:
            query += " AND a.type = %s"
            params.append(type_filter)
        
        if sort_by == 'student_name':
            sort_column = 's.name'
        elif sort_by == 'student_id':
            sort_column = 's.student_id'
        elif sort_by == 'type':
            sort_column = 'a.type'
        else:
            sort_column = 'a.created_at_utc'
        
        query += f" ORDER BY {sort_column} {sort_order.upper()}"
        
        attendance_data = query_all(query, params)
        
        if not attendance_data:
            flash("No data to export for the selected filters")
            return redirect(url_for('admin_panel'))
        
        df = pd.DataFrame(attendance_data)
        df = df.rename(columns={
            'student_id': 'Student ID',
            'student_name': 'Student Name',
            'type': 'Attendance Type',
            'created_at_utc': 'UTC Timestamp',
            'date_local': 'Date',
            'time_local': 'Time'
        })
        
        columns_order = ['Student ID', 'Student Name', 'Attendance Type', 'Date', 'Time', 'UTC Timestamp']
        df = df[columns_order]
        
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Data"
        
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_data_{timestamp}.xlsx"
        
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        flash(f"Error exporting Excel file: {str(e)}")
        return redirect(url_for('admin_panel'))

@app.route("/admin/delete_student/<int:student_id>", methods=["POST"])
@login_required
def delete_student(student_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM students WHERE id=%s", (student_id,))
        conn.commit()
        flash("Student deleted")
    finally:
        conn.close()
    return redirect(url_for("admin_panel"))

@app.route("/admin/delete_token/<token>", methods=["POST"])
@login_required
def delete_token(token):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pending_enrollments WHERE token=%s", (token,))
        conn.commit()
        flash("Token deleted")
    finally:
        conn.close()
    return redirect(url_for("admin_panel"))

@app.route("/admin/delete_all_tokens", methods=["POST"])
@login_required
def delete_all_tokens():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pending_enrollments")
        conn.commit()
        flash("All pending enrollments deleted")
    finally:
        conn.close()
    return redirect(url_for("admin_panel"))

@app.route("/admin/delete_selected_tokens", methods=["POST"])
@login_required
def delete_selected_tokens():
    tokens = request.form.getlist("tokens")
    if not tokens:
        flash("No tokens selected")
        return redirect(url_for("admin_panel"))

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.executemany("DELETE FROM pending_enrollments WHERE token=%s", [(t,) for t in tokens])
        conn.commit()
        flash(f"Deleted {len(tokens)} pending enrollment(s)")
    finally:
        conn.close()
    return redirect(url_for("admin_panel"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

if __name__ == "__main__":
    app.run(debug=False)